import os
import io
import pickle
import platform
import subprocess
import app_context
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient.errors import HttpError
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json
from window_helper import ToplevelWindowHelper
import tkinter as tk
from tkinter import messagebox
import sys
import re
from datetime import date, datetime

class DriveManager:
    """
    A helper class to manage Google Drive operations such as authentication, file download, and file upload.

    Parameters
    ----------
    credentials_file (str, optional): Default = 'credentials.json' 
        Path to the JSON credentials file used for OAuth 2.0 authentication with Google Drive.    
    token_file (str, optional): Default = 'token.json'
        Path to the local JSON token file that caches access credentials to avoid re-authentication 
        on each run.    

    Attributes
    ----------
    credentials_file (str):  
        Path to the JSON credentials file.  
    token_file (str):  
        Path to the token file for caching access credentials.  
    service (googleapiclient.discovery.Resource):  
        Authorized Google Drive API service object, created during initialization 
        and reused for all Drive requests.  

    Methods
    -------
    _authenticate():  
        Authenticate with Google Drive using credentials and return a service object. 
        Handles refreshing and caching tokens automatically.  
    download_file(file_id (str), output_path (str)) -> (bool, str | None):  
        Download a file from Google Drive by its file ID and save it locally at the given path.  
    upload_file(file_id (str), file_path (str)) -> (bool, str | None):  
        Upload (update) a local file to an existing file on Google Drive, specified by its file ID.
    get_archive_dates(folder_id (str)) -> (list):
        Retrieve all archived folder dates from a given parent folder in Google Drive.
    make_archive_copies():
        Creates copies of all the inventory files in a new folder.
    auto_archive():
        Automatically archives the inventory files if there is no archive within the half year.
    """

    # Full access to user's Google Drive
    SCOPES = ['https://www.googleapis.com/auth/drive']

    def __init__(self, credentials_file='credentials.json', token_file='token.json'):
        """ Initialize the Google Drive client helper. See class docstring for parameter/attribute details."""

        self.credentials_file = credentials_file
        self.token_file = token_file
        self.service = self._authenticate()

    def _authenticate(self):
        """
        Handles authentication and returns an authorized Google Drive API service object to access the givne Drive.

        - Loads cached credentials from token_file if available.
        - Refreshes the token if expired.
        - Otherwise runds 0Auth with credentials_file and saves the token for future sessions.
        """

        creds = None

        # Load cached credentials if they exist
        if os.path.exists(self.token_file):
            with open(self.token_file, 'rb') as token:
                creds = pickle.load(token)

        # If no valid credentials, either refresh them or start 0Auth flow
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                # Refresh the expired token
                creds.refresh(Request())
            else:
                # Start a new 0Auth flow. Opens the browser for user to login.
                flow = InstalledAppFlow.from_client_secrets_file(self.credentials_file, self.SCOPES)
                creds = flow.run_local_server(port=0)

            # Save the token locally for reuse next session
            with open(self.token_file, 'wb') as token:
                pickle.dump(creds, token)

        # Build and return the Google Drive API client
        return build('drive', 'v3', credentials=creds)

    def download_file(self, file_id, output_path):
        """
        Download a file from Google Drive using its file ID and save it locally.

        :param file_id (str): The Google Drive file ID to download.
        :param output_path (str): Local path where the file is saved.
        :return: (True, None) on success, or (False, error_message) on fail.
        """

        # Return error if no file_id is present
        if not file_id:
            return False, "MISSING_ID"
        
        try:
            # Create a request to download the file
            request = self.service.files().get_media(fileId=file_id)

            metadata = self.service.files().get(
                fileId=file_id,
                fields="id, name, modifiedTime"
            ).execute()

            last_modified = metadata.get("modifiedTime")

            print(f"⏳ Downloading {output_path} from Google Drive... please wait.")

            # Use a memory buffer to hold the download
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)

            # Download in chunks until complete
            done = False
            while not done:
                status, done = downloader.next_chunk()
                print(f"Download {int(status.progress() * 100)}%")

            # Write the buffer contents to a local file
            with open(output_path, 'wb') as f:
                f.write(fh.getbuffer())

            print(f"✅ Download complete: {output_path}")
            return True, None
        
        # Handle Google Drive API errors by status code
        except HttpError as e:
            if e.resp.status == 404:
                return False, "NOT_FOUND"  # File does not exist
            elif e.resp.status == 403:
                return False, "PERMISSION_DENIED"  # User does not have permission
            else:
                return False, f"HTTP_ERROR: {e}"  # Other errors


    def upload_file(self, file_path, download_alter_time, file_id=None, parent_id=None):
        """
        Upload a file to Google Drive.
        - If file_id is provided, updates the existing file.
        - If file_id is None, creates a new file. If parent_id is given,
        uploads into that folder; otherwise uploads to My Drive.

        :param file_path (str): Local path of the file to upload.
        :param file_id (str): (Optional) Google Drive file ID of an existing file.
        :param parent_id (str): (Optional) Google Drive folder ID to upload into if creating new.
        :return (tuple): 
            (True, file_id) on success
            (False, error_message) on failure.
        """

        try:
            # Create a media upload object for the file to allow Google Drive API to read and upload
            media = MediaFileUpload(
                file_path,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                resumable=True
            )
            # Before update
            metadata = self.service.files().get(
                fileId=file_id,
                fields="id, modifiedTime"
            ).execute()

            current_modified = metadata.get("modifiedTime")
            if current_modified != download_alter_time:
                return False, "STALE_FILE_ERROR" 

            if file_id:
                # Case 1: Update an existing file if a file_id is given
                print(f"⏳ Updating {file_path} on Google Drive... please wait.")
                # Use the Drive API 'update' method to overwrite existing file content
                self.service.files().update(fileId=file_id, media_body=media).execute()
                print(f"✅ File updated on Google Drive (ID: {file_id})")
                return True, file_id

            else:
                # Case 2: Creating a new file if no file_id given
                file_metadata = {
                    # Set the Drive filename same as local filename
                    "name": os.path.basename(file_path)
                }

                if parent_id:
                    # If a parent folder ID is provided, place the file inside that folder
                    file_metadata["parents"] = [parent_id]

                print(f"⏳ Uploading new file {file_path} to Google Drive... please wait.")
                # Use the Drive API 'create' method to create a new file with the metadata and content
                new_file = self.service.files().create(
                    body=file_metadata,
                    media_body=media,
                    fields="id"
                ).execute()

                # Extract the Drive file ID
                new_file_id = new_file.get("id")
                print(f"✅ File uploaded to Google Drive (ID: {new_file_id})")
                return True, new_file_id

        except HttpError as e:
            if e.resp.status == 404:
                return False, "NOT_FOUND"
            elif e.resp.status == 403:
                return False, "PERMISSION_DENIED"
            else:
                return False, f"HTTP_ERROR: {e}"

        except Exception as e:
            return False, f"ERROR: {e}"

    def get_archive_dates(self, folder_id):
        """
        Retrieve all archived folder dates from a given parent folder in Google Drive.

        Looks for subfolders whose names end with a date in teh format MM_DD_YYYY, 
        turns those dates into datetime objects, and returns them as a list.

        :param folder_id (str): The Google Drive ID of the folder to search.
        :return dates (list): A list of the folders dates as datetime objects
        """

        # Query to list only folders inside that parent folder
        query = f"'{folder_id}' in parents and mimeType = 'application/vnd.google-apps.folder' and trashed = false"

        # Execute the API call to list folders
        results = self.service.files().list(
            q=query,
            fields="files(id, name)"
        ).execute()

        folders = results.get("files", [])

        # Regex pattern: Search for MM_DD_YYYY at the end of the string
        pattern = re.compile(r"(\d{2}_\d{2}_\d{4})$")

        dates = []
        for folder in folders:
            # Extract date from folder name if it matches the pattern
            match = pattern.search(folder["name"])
            if match:
                date_str = match.group(1)  # e.g., "07_15_2023"
                # Convert to datetime object if you want
                date_obj = datetime.strptime(date_str, "%m_%d_%Y")
                dates.append(date_obj)

        print("Extracted archive dates:", dates)

        return dates
    
    def make_archive_copies(self):
        """
        Create a new archive folder in Google Drive containing a copy of all the current inventory files.

        This method:
        1. Creates a new folder named "Archive_MM_DD_YYYY" under the main Inventories folder.
        2. Downloads all current inventory files locally.
        3. Uploads each file into the newly created archive folder.
        4. Marks temporary local copies for deletion.
        """

        # Current date formatted as MM_DD_YYYY
        curr_date = date.today().strftime("%m_%d_%Y")

        # Metadata for new archive folder
        file_metadata = {
            "name": f"Archive_{curr_date}",
            "mimeType": "application/vnd.google-apps.folder",
            # Gets the location in Google Drive to place the new folder
            "parents": [app_context.id_manager.get_id("Inventories_Folder")]
        }

        # Create the new archive folder in Drive
        folder = self.service.files().create(body=file_metadata, fields="id").execute()
        # Get the Drive ID of the new folder
        new_folder_id = folder.get("id")
    
        # Loop through all registered inventory files
        for name, file_id in app_context.id_manager.get_all_ids().items():
            # Skip the main folder ID
            if name == "Inventories_Folder":
                continue
            # Create the file name
            curr_path = f"{name}_{curr_date}.xlsx"
            self.download_file(file_id, curr_path)
            # Mark the downloaded file for deletion on program close
            app_context.temp_file_manager.mark_for_deletion(curr_path)
            # Upload file into the new archive folder
            self.upload_file(file_path=curr_path, parent_id=new_folder_id)
            
    def auto_archive(self):
        """
        Automatically create a new archive if no archive exists for the current half-year period.

        Logic:
        1. Determines the current half-year:
        - 1 = January to June
        - 2 = July to December
        2. Retrieves all existing archive folder dates from the main Inventories folder.
        3. Checks if any existing archive falls in the same year and half-year as today.
        4. If no archive exists for the current half, calls make_archive_copies() to create one.
        5. Otherwise, no archive is made.
        """

        # Current date
        current_date = date.today()
        # Fetch existing archive folder dates
        archive_dates = self.get_archive_dates(app_context.id_manager.get_id("Inventories_Folder"))

        # Check if any existing archive is in the same year and half
        given_half = 1 if current_date.month <= 6 else 2
        given_year = current_date.year

        # Check if any other date is in the same year and half
        same_half_exists = any(
            (d.year == given_year) and
            ((1 if d.month <= 6 else 2) == given_half)
            for d in archive_dates
        )
     
        if not same_half_exists:
            # If there is no current archive folder for the half-year, create a new one
            self.make_archive_copies()
            print(f"Created archive copy for {current_date}.")
        else:
            # If there is a current archive folder for the half-year, do nothing
            return
            
class ExcelHelper:
    """
    Manage Excel file associated tasks including opening, creating a DataFrame, Updating and restoring column width.

    Methods
    -------
    open_excel_file(filepath (str)) -> None:  
        Open an Excel file with the system’s default application.  
    create_df(file_path (str), sheet_name (str, optional): Default = 'Sheet1') -> (pd.DataFrame, dict):  
        Load an Excel sheet into a DataFrame. 
    update_single_sheet(file_path (str), df (pd.DataFrame), sheet_name (str)) -> None:  
        Overwrite a single sheet in an Excel file with new DataFrame contents, preserving other sheets.  
    """
    
    @staticmethod
    def open_excel_file(filepath):
        """
        Open an Excel file from the given filepath.

        :param file_path (str): Path to the Excel file.
        """

        system = platform.system()
        if system == "Windows":
            # Open with default app in windows
            os.startfile(filepath)
        elif system == "Darwin":
            # USe 'open' command for macOS
            subprocess.call(["open", filepath])
        elif system == "Linux":
            # USe 'xdg-open' command for linux
            subprocess.call(["xdg-open", filepath])
        else:
            # Error message if OS is not supported
            print(f"Unsupported OS: {system}")
    
    @staticmethod
    def create_df(file_path, sheet_name='Sheet1'):
        """
        Load an Excel sheet into a pandas DataFrame.

        - Converts datetime columns into "mm/dd/yyyy" formatted strings.

        :param file_path (str): Path to the Excel file.
        :param sheet_name (str): Name of the sheet to load. Defaults to 'Sheet1'.
        :return: (DataFrame, dict) -> DataFrame of sheet contents.
        """

        # Load sheet into pandas DataFrame
        df = pd.read_excel(file_path, sheet_name)

        return df

    @staticmethod
    def update_single_sheet(file_path, df, sheet_name):
        """
        Overwrite a single sheet in an Excel file with new DataFrame contents.

        - Creates the workbook if it does not exist.
        - Replaces all rows in the specified sheet with new DataFrame data.
        - Preserves all other sheets

        :param file_path (str): Path to the Excel file.
        :param df (pd.DataFrame): The DataFrame to write into the sheet.
        :param sheet_name (str): Name of the sheet to update. Defaults to 'Sheet1'.
        """

        # Create a new workbook if the file doesn't exist and name the sheet sheet_name
        if not os.path.exists(file_path):
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            wb.create_sheet(title=sheet_name)
            wb.save(file_path)

        wb = load_workbook(file_path)

        # Ensure that the sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name} not found in workbook")

        ws = wb[sheet_name]

        # Delete all rows to clear sheet
        ws.delete_rows(1, ws.max_row)

        # Append DataFrame contents row by row including the header
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        wb.save(file_path)
        wb.close()

class IDManager:
    """
    A helper class to manage Google Drive file IDs used in the application.

    - Loads IDs from a JSON config file (or uses defaults if missing).  
    - Allows updating, resetting, and retrieving IDs.  
    - Provides a Tkinter popup for user-friendly ID correction.  

    Parameters
    ----------
    config_file (str, optional): Default = "drive_ids.json"  
        Path to the JSON file where IDs are stored.  
    default_ids (dict, optional): Default = None  
        Dictionary of default keys/IDs used if the config file does not exist.  

    Attributes
    ----------
    config_file (str):  
        Path to the JSON file where IDs are stored.  
    default_ids (dict):  
        Dictionary of default keys/IDs.  
    _id_registry (dict):  
        Stores the file IDs in memory, loaded from the config file or defaults.  

    Methods
    -------
    get_id(key (str)) -> str | None:  
        Retrieve a stored ID by its key (e.g., "freezer80").  
    update_id(key (str), new_id (str)) -> None:  
        Update a single ID in memory and persist the change to the JSON file.  
    get_all_ids() -> dict:  
        Return a shallow copy of all stored IDs.  
    change_id_window(parent (tk.Widget), key (str)) -> None:  
        Open a Tkinter popup window to update a file ID interactively.  
    """

    def __init__(self, config_file="drive_ids.json", default_ids=None):
        """ Initializes the IDManager. See class dostring for parameter/attribute details"""
        
        if default_ids is None:
            default_ids = {
                "grid_dewar": "",
                "freezer80": "",
                "freezer20": "",
                "cell_culture_rows": "",
                "cell_culture_grid": ""
            }
        self.config_file = config_file
        self.default_ids = default_ids
        # Load IDs from file or initialize with defaults
        self._id_registry = self._load_ids()

    def _load_ids(self):
        """Load IDs from JSON file, or use defaults if file doesn't exist."""
        
        # Get file IDs from config_file and return them
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as f:
                return json.load(f)
        # If no file, save defaults and return them
        else:
            self._save_ids(self.default_ids)
            return self.default_ids.copy()

    def _save_ids(self, ids):
        """Save IDs dictionary to JSON file."""

        with open(self.config_file, "w") as f:
            json.dump(ids, f, indent=4)

    def get_id(self, key):
        """Retrieve an ID by its key (Ex: 'freezer80')."""

        return self._id_registry.get(key)

    def update_id(self, key, new_id):
        """
        Update a single ID in memory and persist to file.
        
        :param key (str): The ID key to update (Ex: 'freezer80').
        :param new_id (str): The new Google Drive file ID.
        """
        self._id_registry[key] = new_id
        self._save_ids(self._id_registry)

    def get_all_ids(self):
        """Return a copy of all stored IDs (dict)."""

        return self._id_registry.copy()
    
    def change_id_window(self, parent, key):
        """
        Display a Tkinter popup to let the user update a missing/invalid Google Drive ID.

        :param parent: The parent Tkinter window.
        :param key (str): The ID key being updated (e.g., 'freezer80').
        """

        # Create a new popup window using ToplevelWindowHelper
        ID_change_helper = ToplevelWindowHelper(parent, "Update Google Drive ID", size="700x500")
        frame = ID_change_helper.get_main_frame()

        # Instruction label
        tk.Label(
            frame,
            text="The Google Drive ID is missing or invalid."
                "Find the file on Google Drive, copy the ID from the URL (between '/d/' and next '/'),"
                "and enter it below:",
            font=("Arial", 14),
            wraplength=500,
            justify="center"
        ).grid(row=0, column=0, padx=40, pady=(20, 30))

        # Show current incorrect ID and its Key
        tk.Label(
            frame,
            text=f"Incorrect ID: {key}\nCurrent ID: {self.get_id(key)}",
            font=("Arial", 12),
            justify="center"
        ).grid(row=1, column=0, padx=20, pady=(10, 20))

        # Entry field (widget) for new ID
        tk.Label(frame, text="Enter New Google Drive ID:", font=("Arial", 12)).grid(row=2, column=0, sticky="w", padx=20)
        id_entry = tk.Entry(frame, font=("Arial", 12), width=50)
        id_entry.grid(row=3, column=0, padx=20, pady=(5, 20))

        # Restart the program after updating
        def restart_program():
            app_context.temp_file_manager.cleanup_temp_files()
            python = sys.executable
            os.execv(python, [python] + sys.argv)

        # Handle update button click
        def on_update():
            new_id = id_entry.get().strip()
            if new_id:
                # Save new ID
                self.update_id(key, new_id)
                print(f"✅ Updated ID for '{key}' to {new_id}")
                messagebox.showinfo("Success", f"Google Drive ID updated for '{key}'.\nPress 'OK' to restart the program.")
                restart_program()
            else:
                messagebox.showerror("Error", "Please enter a valid Google Drive ID.")

        # Update button
        tk.Button(frame, text="Update & Retry", command=on_update, font=("Arial", 12)).grid(
            row=4, column=0, padx=100, pady=(10, 10), sticky="ew"
        )
        frame.grid_columnconfigure(0, weight=1)