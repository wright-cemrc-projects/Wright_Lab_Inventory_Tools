import tkinter as tk
import sys
import os
import pandas as pd
import re
from datetime import datetime
import app_context
from window_helper import ToplevelWindowHelper
from window_configure_helper import dataAddWindows
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import math
from data_helper import ExcelHelper, DriveManager
from tkinter import messagebox
from openpyxl.utils import column_index_from_string
import time

class InventoryManagerBase:
    """
    Base class for managing laboratory inventory stored in Excel/Google Drive.

    Responsibilities
    ----------------
    - Download and upload inventory files (row and grid) from Google Drive.
    - Load Excel data into a cleaned pandas DataFrame.
    - Provide hooks for adding and removing data entries through Tkinter GUIs.
    - Maintain consistent data cleaning/validation rules across the app.
    - Supply default values for configuration (can be overridden in subclasses).

    Typical usage
    -------------
    Subclass this class and override the getter methods
    (e.g., ``get_location_column``, ``get_sort_columns``, ``get_int_fields``).
    Then call ``open_main_menu()`` to show the GUI menu with "View", "Add", 
    and "Remove". The base class handles data cleaning, saving, and 
    Google Drive synchronization automatically.

    Parameters
    ----------
    parent (tk.Widget): 
        Parent Tkinter widget (usually the root window).    
    row_path (str, optional): Default = None
        Local file path to the row inventory Excel file.    
    grid_path (str, optional): Default = None
        Local file path to the grid inventory Excel file (if applicable).    

    Attributes
    ----------
    parent (tk.Widget):  
        Parent Tkinter window.  
    row_path (str):  
        Path to the local copy of the row inventory file.  
    grid_path (str):  
        Path to the local copy of the grid inventory file.  
    row_ID (str):  
        Google Drive file ID for the row inventory.  
    grid_ID (str):  
        Google Drive file ID for the grid inventory.  
    row_id_key (str):  
        Key used by the ID manager for the row file.  
    grid_id_key (str):  
        Key used by the ID manager for the grid file.  
    rows_df (pd.DataFrame):  
        DataFrame holding the row inventory data.   

    Methods
    -------
    clean_dataframe(df (pd.DataFrame)) -> pd.DataFrame:  
        Return a cleaned copy of the DataFrame with normalized types and values.  
    load_data():  
        Download inventory Excel files from Google Drive, load into DataFrames,
        and apply cleaning rules.  
    get_row_sheet_name() -> str:  
        Return the default Excel sheet name for row inventory data.  
    open_main_menu():  
        Open a Tkinter window with options to view, add, or remove inventory entries.  
    click_view_inventory():  
        Open the current inventory Excel file(s) in the system’s default program.  
    handle_add(updated_rows_df (pd.DataFrame)):  
        Callback to handle new rows added via the GUI and save changes.  
    handle_remove(updated_rows_df (pd.DataFrame)):  
        Callback to handle row removals via the GUI and save changes.  
    update_row_inventory():  
        Save the row inventory Excel file and upload it to Google Drive.  
    update_grid_inventory():  
        Placeholder to update grid inventory; override in subclasses if applicable.  
    Changes_Saved_Window():  
        Display a confirmation window after changes have been saved.  

    Getter methods (override in subclasses)
    ---------------------------------------
    get_window_title() -> str:  
        Return the title for the main inventory window.  
    get_sort_columns() -> list[str]:  
        Return columns to sort the inventory DataFrame by.  
    get_for_filled() -> list[str]:  
        Return columns used to determine if a location is occupied.  
    get_location_column() -> str:  
        Return the column representing inventory locations.  
    get_int_fields() -> list[str]:  
        Return fields that should be interpreted as integers.  
    get_remove_fields() -> list[str]:  
        Return columns used to identify rows for removal.  
    get_label_column() -> str:  
        Return the column used as a display label.  
    get_letterNums() -> list[str]:  
        Return fields where letter-number parsing is required.  
    get_date_column() -> str | None:  
        Return the column storing date values.  
    get_grid_coords() -> tuple[int, int] | None:  
        Return grid dimensions (rows, columns) for picker layout.  
    get_rectangle_picker() -> bool:  
        Return True if the grid picker should be rectangular.  
    get_required_add_fields() -> list[str]:  
        Return fields required when adding a new entry.  
    get_add_top_label() -> str:  
        Return the label displayed at the top of the Add Entry window.  
    get_add_window_name() -> str:  
        Return the name of the Add Entry window.  
    get_required_remove_fields() -> list[str]:  
        Return fields required when removing an entry.  
    get_remove_top_label() -> str:  
        Return the label displayed at the top of the Remove Entry window.  
    get_remove_window_name() -> str:  
        Return the name of the Remove Entry window.  
    get_unused_columns() -> list[str]:  
        Return columns to hide in the GUI.  
    """

    def __init__(self, parent, row_path=None, grid_path=None):
        """Initialize the Inventory Manager. See class docstring for parameter/attribute details."""

        self.parent = parent
        self.row_path = row_path
        self.grid_path = grid_path
        self.row_ID = None
        self.grid_ID = None
        self.row_id_key = None
        self.grid_id_key = None
        self.rows_df = None
    
    def clean_dataframe(self, df):
        """Return a cleaned copy of the DataFrame.
        
        This ensures consistent formatting for integers, dates, letter-number fields, 
        and missing values across the entire DataFrame. The cleaning rules come 
        from `clean_df_value`, applied column by column. Uses clean_df_value() to clean each value.

        :param df (pd.DataFrame): The data frame to be cleaned.
        :return df_cleaned (pd.DataFrame): The data frame after its values are cleaned.
        """

        # Work on a copy of the DataFrame
        df_cleaned = df.copy()

        # Cache column types
        int_cols = set(self.get_int_fields())
        letter_num_cols = set(self.get_letter_nums())
        date_col = self.get_date_column()
        all_cols = set(df_cleaned.columns)
        other_cols = all_cols - int_cols - letter_num_cols - {date_col}

        # --- Integer columns ---
        for col in int_cols:
            if col in df_cleaned.columns:
                df_cleaned[col] = pd.to_numeric(df_cleaned[col], errors="coerce").astype("Int64")

        # --- Letter-number columns (e.g., A12) ---
        pattern = r"^[A-Za-z]{1}\d+$"
        for col in letter_num_cols:
            if col in df_cleaned.columns:
                # Convert to string, strip whitespace, uppercase
                series = df_cleaned[col].astype(str).str.strip().str.upper()
                # Keep only values matching the pattern, else NaN
                df_cleaned[col] = series.where(series.str.match(pattern))

        # --- Date column ---
        if date_col and date_col in df_cleaned.columns:
            # Try converting values to datetime to capture dates with invalid format (invalid entries become NaT)
            series = pd.to_datetime(df_cleaned[date_col], errors="coerce", dayfirst=False)
            # Format the dates as MM/DD/YYYY (invalid become NaN)
            df_cleaned[date_col] = series.dt.strftime("%m/%d/%Y")
            # Format invalid into empty string ""
            df_cleaned[date_col] = df_cleaned[date_col].fillna("")

        # --- 4 All other columns (string/text) ---
        for col in other_cols:
            if col in df_cleaned.columns:
                # Convert to string, strip whitespace, uppercase
                df_cleaned[col] = df_cleaned[col].astype(str).str.strip()
                df_cleaned[col] = df_cleaned[col].replace({"nan": ""})

        # Fill in missing values consistently
        for col in df_cleaned.columns:
            if col in self.get_int_fields():
                # -1 for int columns
                df_cleaned[col] = df_cleaned[col].fillna(-1)
            else:
                # Empty string for all other columns
                df_cleaned[col] = df_cleaned[col].fillna("")

        # Return the fully cleaned DataFrame
        return df_cleaned

    def _handle_drive_error(self, error, id_key):
        # Handle different error cases from DriveManager
        if error =="STALE_FILE_ERROR":
            messagebox.showerror("The file you are attempting to update had been modified since your download. Please close the program and restart your action to continue.")
        elif error == "MISSING_ID" or error == "NOT_FOUND":
            # Prompt user to update ID if file is missing
            app_context.id_manager.change_id_window(self.parent, id_key)
        elif error == "PERMISSION_DENIED":
            # User doesn’t have Drive access to this file
            messagebox.showerror("Permission Denied", "You don't have access to this file.")
        else:
            # Any other unexpected error
            messagebox.showerror("Download Failed", f"Unexpected error: {error}")  
        

    def load_data(self):
        """
        Download inventory Excel files from Google Drive, load into DataFrames, and clean.

        - Always loads the "row" inventory file.
        - Optionally loads a "grid" inventory file if IDs and paths are provided.
        - Handles Google Drive errors (missing ID, not found, permission issues).
        - Marks downloaded files for later deletion by the temp file manager.
        - Saves a variable of when the files where last altered
        - Applies cleaning rules to the loaded DataFrame(s) by clean_dataframe().
        """

        self.drive_tool = DriveManager()
        self.drive_tool.auto_archive()

        # --- Step 1: Download and load row data ---
        if not (self.row_id_key == None):
            success, info = self.drive_tool.download_file(self.row_ID, self.row_path)
            if not success:
                self._handle_drive_error(info, self.row_id_key)
            if success:
                # Store the last time the row file was altered on Google Drive
                self.row_alter_time = info

            # Ensure this temp file is tracked for cleanup at exit         
            app_context.temp_file_manager.mark_for_deletion(self.row_path)

            # Read Excel sheet into DataFrame
            self.rows_df = ExcelHelper.create_df(self.row_path, sheet_name=self.get_row_sheet_name())
            # Apply cleaning rules
            self.rows_df = self.clean_dataframe(self.rows_df) 

        # --- Step 2: Download and load grid data ---
        if self.row_ID == self.grid_ID:
            # Skip if it is in the same file as the rows data (even if on a seperate sheet)
            return
        if not (self.grid_id_key == None):
            success, error = self.drive_tool.download_file(self.grid_ID, self.grid_path)
            if not success:
                self._handle_drive_error(error, self.grid_id_key)
            if success:
                # Store the last time the grid file was altered on Google Drive
                self.grid_alter_time = info

            # Ensure this temp file is tracked for cleanup at exit
            app_context.temp_file_manager.mark_for_deletion(self.grid_path)

    def get_row_sheet_name(self):
        """
        Return the default sheet name in the row Excel file.

        Subclasses can override this method to specify a different sheet
        name if the row inventory data is not stored in "Sheet1".
        
        :return (str): Name of the worksheet containing row inventory data.
        """

        return "Sheet1"

    def open_main_menu(self):
        """
        Build and display the main menu window for an inventory manager.

        This method:
        - Loads inventory data from Google Drive / Excel into a cleaned DataFrame.
        - Creates a new Tkinter toplevel window with a standard header and buttons.
        - Instantiates a `dataAddWindows` manager to handle add/remove workflows.
        - Provides the user with three primary options:
            1. View current inventory in Excel.
            2. Add new entries to the inventory.
            3. Remove entries from the inventory.
        - Connects add/remove callbacks so that updates are propagated back to 
        the DataFrame and saved to Drive.
        """

        # Load data from Drive/Excel into cleaned DataFrame
        self.load_data()

        # Create a toplevel window for the main menu
        window_helper = ToplevelWindowHelper(self.parent, self.get_window_title())
        main_frame = window_helper.get_main_frame()
        window = window_helper.window

        # Initialize dataAddWindows manager 
        windowManager = dataAddWindows(
            parent=window,
            rows_df=self.rows_df,
            columns_to_sort_by=self.get_sort_columns(),
            for_filled=self.get_for_filled(),
            location_column=self.get_location_column(),
            int_fields=self.get_int_fields(),
            remove_fields=self.get_remove_fields(),
            label_column=self.get_label_column(),
            letterNums=self.get_letter_nums(),
            date_column=self.get_date_column(),
            grid_coords=self.get_grid_coords(),
            rectangle_picker=self.get_rectangle_picker()
        )

        # Hook callbacks so add/remove actions trigger inventory updates
        windowManager.add_callback = self.handle_add
        windowManager.remove_callback = self.handle_remove

        # Create header label
        label = tk.Label(main_frame, text="What do you want to do?", font=("Arial", 16))
        label.grid(row=1, column=0, columnspan=3, sticky="ew", padx=50, pady=(10, 0))

        # Define the main menu buttons
        buttons = [
            ("View Current Inventory", self.click_view_inventory),
            ("Add to Inventory", lambda: (app_context.temp_file_manager.notify_if_open_files(), windowManager.Configure_AddRemove_Window(
                required_fields=self.get_required_add_fields(),
                top_label=self.get_add_top_label(),
                adding=True,
                unused_columns=self.get_unused_columns(),
                window_name=self.get_add_window_name(),
            ))),
            ("Remove from Inventory", lambda: (app_context.temp_file_manager.notify_if_open_files(), windowManager.Configure_AddRemove_Window(
                required_fields=self.get_required_remove_fields(),
                top_label=self.get_remove_top_label(),
                adding=False,
                unused_columns=self.get_unused_columns(),
                window_name=self.get_remove_window_name(),
            ))),
        ]

        # Place buttons on the window in a grid layout
        for i, (text, cmd) in enumerate(buttons):
            btn = tk.Button(main_frame, text=text, command=cmd, font=("Arial", 14),
                            relief="groove", activebackground="gray")
            btn.grid(row=(i+2), column=0, columnspan=3, sticky="nsew", padx=50, pady=(20, 20))

        # Configure equal expansion for rows/columns
        for col in range(3):
            main_frame.grid_columnconfigure(col, weight=1)
        for i in range(1, len(buttons) + 2):
            main_frame.grid_rowconfigure(i, weight=1)

    def click_view_inventory(self):
        """
        Open the current inventory Excel file(s) for direct viewing by the user.

        - Always opens the "row" inventory file (self.row_path).
        - If a grid-based inventory file is associated with this manager,
        it will be opened as well (self.grid_path).
        
        Notes
        -----
        This uses ExcelHelper.open_excel_file(), which launches the file 
        in the system’s default program (usually Excel).
        """

        # Open row inventory file
        ExcelHelper.open_excel_file(self.row_path)
        # Open grid inventory file if available
        if self.grid_path:
            ExcelHelper.open_excel_file(self.grid_path)

    def handle_add(self, updated_rows_df):
        """
        Handle the addition of new rows to the inventory.

        - This method is called as a callback after new data has been 
          entered and validated in the "Add to Inventory" workflow.
        - Updates the rows dataframe, grid layout, and then uploads the necessary files to Google Drive.

        :param updated_rows_df (pd.DataFrame): The updated row inventory data after new entries have been added.
        """

        # Replace the current rows_df with the updated version from the callback
        self.rows_df = pd.DataFrame(updated_rows_df)
        # Save updates to row inventory in Google Drive
        self.update_row_inventory()
        # Save updates to grid inventory in Google Drive
        self.update_grid_inventory()

        # Upload the updated row file to Google Drive
        success, info = self.drive_tool.upload_file(file_path=self.row_path, file_id=self.row_ID, download_alter_time=self.row_alter_time)

        # If upload fails, handle the error gracefully
        if not success:
            self._handle_drive_error(info, self.row_id_key)
        
        if self.grid_ID != self.row_ID:
            # Upload the updated grid file to Google Drive if different from row file
            success, info = self.drive_tool.upload_file(file_path=self.grid_path, file_id=self.grid_ID, download_alter_time=self.grid_alter_time)

            # If upload fails, handle the error gracefully
            if not success:
                self._handle_drive_error(info, self.row_id_key)

        # Show confirmation window that changes have been saved
        self.Changes_Saved_Window()

    def handle_remove(self, updated_rows_df):
        """
        Handle the addition of new rows to the inventory.

        This method is called as a callback after new data has been 
        entered and validated in the "Add to Inventory" workflow.

        :param updated_rows_df (pd.DataFrame): The updated row inventory data after new entries have been added.
        """

        # Replace the current rows_df with the updated version from the callback
        self.rows_df = pd.DataFrame(updated_rows_df)
        # Save updates to row inventory in Google Drive
        self.update_row_inventory()
        # Save updates to grid inventory in Google Drive
        self.update_grid_inventory()

                # Upload the updated row file to Google Drive
        success, error = self.drive_tool.upload_file(file_path=self.row_path, file_id=self.row_ID, download_alter_time=self.row_alter_time)

        # If upload fails, handle the error gracefully
        if not success:
            self._handle_drive_error(error, self.row_id_key)
        
        if self.grid_ID != self.row_ID:
            # Upload the updated grid file to Google Drive if different from row file
            success, error = self.drive_tool.upload_file(file_path=self.grid_path, file_id=self.grid_ID, download_alter_time=self.grid_alter_time)

            # If upload fails, handle the error gracefully
            if not success:
                self._handle_drive_error(error, self.row_id_key)

        # Show confirmation window that changes have been saved
        self.Changes_Saved_Window()

    def update_row_inventory(self):
        """Update the row inventory Excel file and upload to Google Drive."""

        # Write the updated DataFrame back into the Excel sheet
        ExcelHelper.update_single_sheet(self.row_path, self.rows_df, self.get_row_sheet_name())

    def update_grid_inventory(self):
        """
        Update the grid inventory Excel file and synchronize with Google Drive.

        This base implementation is a placeholder.  
        Subclasses should override this method if their inventory type
        includes a separate "grid" file in addition to the row inventory.

        Notes
        -----
        - If the inventory type does not use a grid file, this method
        can remain unimplemented.
        """

        # No default behavior; meant to be overridden in subclasses
        pass

    def Changes_Saved_Window(self):
        """
        Display a confirmation window after changes have been saved.

        This method creates a new top-level window informing the user that
        their changes have been successfully saved. It provides a message
        and a button allowing the user to restart the program if they want
        to continue entering more data.
        """

        # Create a new top-level confirmation window with fixed size
        saved_win_helper = ToplevelWindowHelper(self.parent, self.get_window_title(), size="700x400", )
        saved_win = saved_win_helper.get_main_frame()
        saved_win.grid_columnconfigure(0, weight=1)

        # Add a label with confirmation text and instructions
        label = tk.Label(
            saved_win, text="Your changes have been saved! \nIf you would like to enter more data click the button below, Otherwise you may exit.",
            font=("Arial", 12),
            justify="center")
        label.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="n")

        def restart_program():
            """
            Restart the entire program.

            This cleans up any temporary files created during the session,
            then re-executes the Python interpreter with the original
            command-line arguments (`sys.argv`).
            """
            
            app_context.temp_file_manager.cleanup_temp_files()
            python = sys.executable
            os.execv(python, [python] + sys.argv)

        # Create a button to let the user restart the program
        button = tk.Button(saved_win, text="I have more to enter", command=restart_program, font=("Arial", 12))
        button.grid(row=1, column=0, padx=20, pady=(20, 10))

    # The following are getter methods for subclass override

    def get_window_title(self):
        """
        Return the title of the main inventory management window.

        Override in subclass to provide a more specific title.
        """
        return "Inventory Manager"

    def get_sort_columns(self):
        """
        Return a list of column names to use for sorting the DataFrame. 
        
        - Sorts in order of column names within the List

        Override in subclass to define which columns determine row order.
        """
        return []

    def get_for_filled(self):
        """
        Return a list of columns used to determine whether a location is considered 'filled' or occupied.

        - This excludes the final location column of an individual unit of invnetory (Ex: Vial Location)

        Override in subclass to specify the occupancy check.
        """
        return []

    def get_location_column(self):
        """
        Return the name of the column that represents the location of an individual unit of inventory.

        Override in subclass if inventory has a location field (e.g., 'Rack').
        """
        return ""

    def get_int_fields(self):
        """
        Return a list of fields that should be interpreted strictly as integers.

        Override in subclass to enforce integer validation on certain fields.
        """
        return []

    def get_remove_fields(self):
        """
        Return a list of fields used to identify and remove an inventory entry.

        - Often is a list that contains both get_for_filled and get_location_column

        Override in subclass to define the removal criteria.
        """
        return []

    def get_label_column(self):
        """
        Return the name of the column containing the label of an individual unit of inventory.

        Override in subclass to specify a display label (e.g., 'Sample Name').
        """
        return ""

    def get_letter_nums(self):
        """
        Return a list of fields that should contian a letter followed by a number.

        Override in subclass for grid-based storage systems.
        """
        return []

    def get_date_column(self):
        """
        Return the name of the column that stores date values.

        Override in subclass if inventory entries include dates,
        otherwise return None.
        """
        return None

    def get_grid_coords(self):
        """
        Return the grid dimensions (rows, columns) as a tuple.

        - Return "Check" if the dimensions should be later determined by the DataFrame

        Override in subclass to define the shape of the vial picker grid.
        """
        return None

    def get_rectangle_picker(self):
        """
        Return True if the inventory picker should use rectangular selection,
        False otherwise.

        Override in subclass if using a non-rectangular layout.
        """
        return True

    def get_required_add_fields(self):
        """
        Return a list of fields required when adding a new entry.

        Override in subclass to enforce input requirements.
        """
        return []

    def get_add_top_label(self):
        """
        Return the label text displayed at the top of the 'Add Entry' window.

        Override in subclass to provide context-specific instructions.
        """
        return "Enter Data Below"

    def get_add_window_name(self):
        """
        Return the name of the 'Add Entry' window.

        Override in subclass to provide a more descriptive name.
        """
        return "Add Entry"

    def get_required_remove_fields(self):
        """
        Return a list of fields required when removing an entry.

        Override in subclass to enforce which fields must be specified.
        """
        return []

    def get_remove_top_label(self):
        """
        Return the label text displayed at the top of the 'Remove Entry' window.

        Override in subclass to provide context-specific instructions.
        """
        return "Remove Data Below"

    def get_remove_window_name(self):
        """
        Return the name of the 'Remove Entry' window.

        Override in subclass to provide a more descriptive name.
        """
        return "Remove Entry"

    def get_unused_columns(self):
        """
        Return a list of column names that should be ignored in the GUI.

        Override in subclass to hide columns that are not relevant for display
        or user input.
        """
        return []
    
class GridDewarManager(InventoryManagerBase):
    """
    Manages inventory records for the Grid Dewar system.

    This subclass of InventoryManagerBase provides configuration details
    specific to grid dewar storage, including file paths, required fields,
    sorting rules, and user interface labels. It defines how rows should
    be validated, displayed, and managed for entries in the grid dewar
    inventory.

    Parameters
    ----------
    root (tk.Widget):
        The Tkinter root window or parent widget.

    Attributes
    ----------
    row_id_key (str):
        Key used to look up the file ID for row-based inventory in the ID manager.
    grid_id_key (str or None):
        Key used to look up the file ID for grid-based inventory. Set to None 
        since the grid is not used in this manager.
    row_ID (str):
        File ID for the row-based inventory Excel file.
    grid_ID (str or None):
        File ID for grid-based inventory. Set to None in this manager.

    Methods
    -------
    get_window_title() -> str:
        Return the title of the main inventory window.
    get_row_sheet_name() -> str:
        Return the name of the Excel sheet used for row-based inventory.
    get_sort_columns() -> list[str]:
        Return the columns used to sort the grid dewar inventory.
    get_for_filled() -> list[str]:
        Return the columns used to determine if an inventory position is filled.
    get_location_column() -> str:
        Return the column representing the physical location of an inventory unit.
    get_int_fields() -> list[str]:
        Return the fields that should be strictly treated as integers.
    get_remove_fields() -> list[str]:
        Return the fields required to identify and remove an inventory entry.
    get_label_column() -> str:
        Return the column used as a label for an inventory entry.
    get_letter_nums() -> list[str]:
        Return fields that contain a letter followed by a number (e.g., A4).
    get_date_column() -> str:
        Return the column that stores date information.
    get_rectangle_picker() -> bool:
        Return False if the inventory picker does not use rectangular layouts.
    get_required_add_fields() -> list[str]:
        Return the fields required when adding a new entry.
    get_add_top_label() -> str:
        Return the label shown at the top of the 'Add Entry' window.
    get_add_window_name() -> str:
        Return the name of the 'Add Entry' window.
    get_required_remove_fields() -> list[str]:
        Return the fields required when removing an entry.
    get_remove_top_label() -> str:
        Return the label shown at the top of the 'Remove Entry' window.
    get_remove_window_name() -> str:
        Return the name of the 'Remove Entry' window.
    """

    def __init__(self, root):
        """Initializes the GridDewarManager. See class docstring for parameter/attribute details."""

        # Initialize logic from the InventoryManagerBase using the given row_path
        super().__init__(root, 
            row_path="Grid_Dewar_Inventory.xlsx", 
            grid_path=None
            )
        # Key for row inventory lookup in the ID manager
        self.row_id_key = "Grid_Dewar_Inventory"
        # Grid is not used for Grid Dewar, so this remains None
        self.grid_id_key = None
        # Get the unique ID for the row file from the ID manager
        self.row_ID = app_context.id_manager.get_id(self.row_id_key)
        # No grid file is managed here, so leave None
        self.grid_ID=None

    def get_window_title(self):
        """Return the title for the main window."""

        return "Grid Dewar"

    def get_row_sheet_name(self):
        """Return the name of the Excel sheet used for row-based inventory."""

        return "Details"

    def get_sort_columns(self):
        """Return the list of columns used for sorting the grid dewar inventory in sorting order."""

        return ["Cane Number", "Puck Number", "Slot Number"]

    def get_for_filled(self):
        """
        Return the columns used to determine if an invenotry position is filled.

        This excludes the column containing the location of an individual unit of inventory.
        """

        return ["Cane Number", "Puck Number"]

    def get_location_column(self):
        """Return the column representing the physical location of an individual unit of inventory."""
        
        return "Slot Number"

    def get_int_fields(self):
        """Return the fields that should be strictly treated as integers."""

        return ["Cane Number", "Puck Number", "Slot Number"]

    def get_remove_fields(self):
        """Return the fields required to identify and remove an entry."""

        return ["Cane Number", "Puck Number", "Slot Number"]

    def get_label_column(self):
        """Return the column name used as a label for an individual inventory entry."""

        return "Box Name"

    def get_letter_nums(self):
        """Return the fields that conatin a letter followed by a number (Ex: A4)"""

        return []

    def get_date_column(self):
        """Return the column that stores date information."""

        return "Date Frozen"

    def get_rectangle_picker(self):
        """
        Return False since the grid dewar does not use rectangular picker layouts.
        
        The grid dewar system uses its own picker with a different layout.
        """
        
        return False

    def get_required_add_fields(self):
        """Return the fields required to be enetered by the user when adding a new entry."""

        return ["Cane Number", "Puck Number", "Slot Number", "Box Name", "Grid Numbers", "Box Contents", "Date Frozen", "Person/Initials", "Project", "Grid Type", "Blot Time", "Blot Force" "Drain Time"]

    def get_add_top_label(self):
        """Return the label shown at the top of the 'Add Entry' window."""

        return "Enter Grid Box Information Below"

    def get_add_window_name(self):
        """Return the name of the 'Add Entry' window."""

        return "Grid Dewar Box Entry"

    def get_required_remove_fields(self):
        """Return the fields to be enterd by the user required when removing an entry."""

        return ["Cane Number", "Puck Number", "Slot Number"]

    def get_remove_top_label(self):
        """Return the label shown at the top of the 'Remove Entry' window."""

        return "Enter Grid Box Information Below"

    def get_remove_window_name(self):
        """Return the name of the 'Remove Entry' window."""

        return "Grid Dewar Box Removal"
    
    def get_unused_columns(self):
        """Return column names that are ignored in the GUI."""

        return ['Unnamed: 6']

class Freezer80Manager(InventoryManagerBase):
    """
    Manages inventory records for the -80 Freezer system.

    This subclass of InventoryManagerBase provides configuration details
    specific to -80 Freezer storage, including file paths, required fields,
    sorting rules, and user interface labels. It defines how rows should
    be validated, displayed, and managed for entries in the -80 Freezer
    inventory, and handles updating the Excel-based grid layout with
    colored indicators for person/project assignments.

    Parameters
    ----------
    root (tk.Widget):
        The Tkinter root window or parent widget.

    Attributes
    ----------
    row_id_key (str):
        Key used to look up the file ID for row-based inventory in the ID manager.
    grid_id_key (str or None):
        Key used to look up the file ID for grid-based inventory in the ID manager.
    row_ID (str):
        File ID for the row-based inventory Excel file.
    grid_ID (str or None):
        File ID for grid-based inventory Excel file.

    Methods
    -------
    get_window_title() -> str:
        Return the title of the main inventory window.
    get_row_sheet_name() -> str:
        Return the name of the Excel sheet used for row-based inventory.
    get_sort_columns() -> list[str]:
        Return the columns used to sort the -80 Freezer inventory.
    get_for_filled() -> list[str]:
        Return the columns used to determine if an inventory position is filled.
    get_location_column() -> str:
        Return the column representing the physical location of an inventory unit.
    get_int_fields() -> list[str]:
        Return the fields that should be strictly treated as integers.
    get_remove_fields() -> list[str]:
        Return the fields required to identify and remove an inventory entry.
    get_label_column() -> str:
        Return the column used as a label for an inventory entry.
    get_letter_nums() -> list[str]:
        Return the fields that contain a letter followed by a number (e.g., A4).
    get_date_column() -> str:
        Return the column that stores date information.
    get_grid_coords() -> str:
        Return "Check" to indicate grid dimensions should be read from the DataFrame.
    get_required_add_fields() -> list[str]:
        Return the fields required when adding a new entry.
    get_add_top_label() -> str:
        Return the label shown at the top of the 'Add Entry' window.
    get_add_window_name() -> str:
        Return the name of the 'Add Entry' window.
    get_required_remove_fields() -> list[str]:
        Return the fields required when removing an entry.
    get_remove_top_label() -> str:
        Return the label shown at the top of the 'Remove Entry' window.
    get_remove_window_name() -> str:
        Return the name of the 'Remove Entry' window.
    get_unused_columns() -> list[str]:
        Return column names that should be ignored in the GUI.
    update_grid_inventory() -> None:
        Update the Excel grid layout for the -80 Freezer, clearing existing values,
        placing boxes in the correct shelf/rack/position, assigning colors to persons/projects,
        creating a legend, saving the workbook, and uploading it to Google Drive.
    """

    def __init__(self, root):
        """Initializes the Freezer80Manager. See class docstring for parameter/attribute details."""

        # Initialize logic from the InventoryManagerBase using the given row_path and grid_path
        super().__init__(
            root,
            row_path="80_Inventory.xlsx",
            grid_path="80_Inventory.xlsx"
        )
        # Key for row inventory lookup in the ID manager
        self.row_id_key = "80_Inventory"
        # Key for grid inventory lookup in the ID manager
        self.grid_id_key = "80_Inventory"
        # Get the unique ID for the row file from the ID manager
        self.row_ID = app_context.id_manager.get_id(self.row_id_key)
        # Get the unique ID for the grid file from the ID manager
        self.grid_ID = app_context.id_manager.get_id(self.grid_id_key)

    def get_window_title(self):
        """Return the title for the main window."""

        return "-80 Freezer Dewar"

    def get_row_sheet_name(self):
        """Return the name of the Excel sheet used for row-based inventory."""

        return "Details"

    def get_sort_columns(self):
        """Return the list of columns used for sorting the grid dewar inventory in sorting order."""

        return ["Shelf Number", "Rack Number", "Box Position", "Vial Position"]

    def get_for_filled(self):
        """
        Return the columns used to determine if an inventory position is filled.

        This excludes the column containing the location of an individual unit of inventory.
        """

        return ["Shelf Number", "Rack Number", "Box Position"]

    def get_location_column(self):
        """Return the column representing the physical location of an individual unit of inventory."""

        return "Vial Position"

    def get_int_fields(self):
        """Return the fields that should be strictly treated as integers."""

        return ["Shelf Number", "Rack Number"]

    def get_remove_fields(self):
        """Return the fields required to identify and remove an entry."""

        return ["Shelf Number", "Rack Number", "Box Position", "Box Name", "Vial Position"]

    def get_label_column(self):
        """Return the column name used as a label for an individual inventory entry."""

        return "Vial Label"

    def get_letter_nums(self):
        """Return the fields that conatin a letter followed by a number (Ex: A4)"""

        return ["Box Position", "Vial Position"]
    
    def get_date_column(self):
        """Return the column that stores date information."""

        return "Date Frozen"

    def get_grid_coords(self):
        """Return "Check" for the picker grid coordinates to require the picker to look in the DataFrame for dimensions."""
        
        return "Check"

    def get_required_add_fields(self):
        """Return the fields required to be enetered by the user when adding a new entry."""

        return ["Shelf Number", "Rack Number", "Box Position", "Box Name", "Vial Position", "Vial Label", "Vial Contents", "Date Frozen", "Person/Initials", "Project", "Box Dimensions"]

    def get_add_top_label(self):
        """Return the label shown at the top of the 'Add Entry' window."""

        return "Enter Box and Vial Information Below"

    def get_add_window_name(self):
        """Return the name of the 'Add Entry' window."""

        return "-80 Freezer Vial Data Entry"

    def get_required_remove_fields(self):
        """Return the fields to be enterd by the user required when removing an entry."""

        return ["Shelf Number", "Rack Number", "Box Position", "Box Name", "Vial Position"]

    def get_remove_top_label(self):
        """Return the label shown at the top of the 'Remove Entry' window."""
        
        return "Enter Vial Information Below"

    def get_remove_window_name(self):
        """Return the name of the 'Remove Entry' window."""

        return "-80 Freezer Vial Removal"

    def get_unused_columns(self):
        """Return column names that are ignored in the GUI."""

        return ['Unnamed: 6']

    def update_grid_inventory(self):
        """
        Update the Excel grid inventory layout for the -80°C freezer.

        This method clears existing entries in the predefined grid layout
        and fills in the updated box names from `self.rows_df`. Each box
        is positioned based on its shelf, rack, and position information. Cells
        are automatically colored based on the person or project responsible,
        with a color legend generated on the side of the sheet.

        Steps performed:
        1. Clear all values, formatting, and alignment in the defined row/column blocks.
        2. Define a pool of colors for assigning to persons or projects.
        3. Assign colors dynamically to unique persons/projects as boxes are added.
        4. Validate and place each box in the correct shelf/rack/position.
        5. Create a legend on the side showing assigned colors for persons/projects.
        6. Save the Excel workbook.

        Notes
        -----
        - Expects the Excel workbook to have a "Racks" sheet.
        - Shelf numbers should be 1–5, racks 1–4, row letters A–D, columns 1–5.
        - Any invalid or missing positions are skipped.
        - Fallback color "D3D3D3" is used if color pool is exhausted.

        Methods
        -------
        def get_next_color():
            Return the next unused color from the color pool.
        get_fill_color(row):
            Determine cell fill color based on 'Person/Initials' or 'Project/Group'.
        """

        # Load the workbook and the "Racks" sheet
        wb = load_workbook(self.grid_path)
        sheet = wb["Racks"]

        # Define the row blocks to clear before updating
        row_blocks = [
            range(3, 7),    # Rows 3–6
            range(11, 15),  # Rows 11–14
            range(19, 23),  # Rows 19–22
            range(27, 31),  # Rows 27–30
            range(35, 39),  # Rows 35–38
        ]

        # Define the column groups to clear before updating
        col_groups = [
            ("C", "G"),   # Columns C–G
            ("J", "N"),   # Column J–N
            ("Q", "U"),   # Column Q–U
            ("X", "AB"),  # Column X–AB
        ]

        # Clear existing values, fills, and alignments
        for row_block in row_blocks:
            for start_col, end_col in col_groups:
                start_idx = column_index_from_string(start_col)
                end_idx = column_index_from_string(end_col)
                for r in row_block:
                    for c in range(start_idx, end_idx + 1):
                        cell = sheet.cell(row=r, column=c)
                        cell.value = ""  # clear value
                        cell.fill = PatternFill()  # clear fill
                        cell.alignment = Alignment(
                            wrap_text=True, horizontal="center", vertical="center"
                        )

        # Define a pool of colors for person/project assignments
        color_pool = [
            "FFB3BA", "FFB74D", "FFFFBA", "A5D6A7", "81D4FA", "B39DDB",
            "E53935", "FB8C00", "FDD835", "43A047", "039BE5", "00ACC1", "8E24AA",
            "D7CCC8", "FFE0B2", "8E735B", "6D4C41", "8D6E63",
            "BDBDBD", "757575", "90A4AE", "A5A58D", "546E7A",
            "00897B", "7CB342", "CE93D8", "AED581", "FFD54F", "FF8A65", "C0CA33",
            "FF1744", "D500F9", "00E5FF", "1DE9B6", "FFD600", "FF6D00", "DD2C00",
            "304FFE", "00BFA5", "76FF03", "C51162", "6200EA", "2962FF", "00B0FF",
            "00C853", "FFD600", "FFAB00", "FF3D00", "B71C1C", "1B5E20", "0D47A1",
            "311B92", "F57F17", "FF1744", "D50000", "C2185B", "4A148C", "0091EA",
            "00BFA5", "64DD17", "FFD600"
        ]

        # Track which colors are assigned to which persons/projects
        assigned_person_colors = {}
        assigned_project_colors = {}
        used_colors = set()

        def get_next_color():
            """Return the next unused color from the color pool."""

            for color in color_pool:
                if color not in used_colors:
                    used_colors.add(color)
                    return color
            return "D3D3D3"  # fallback gray

        def get_fill_color(row):
            """
            Determine cell fill color based on 'Person/Initials' or 'Project/Group'.
            
            Priority is given to person. If neither is defined, a default gray is used.
            """
            
            # Get the 'Person/Initials" and "Project/Group" of the row
            person = str(row.get('Person/Initials', '')).strip()
            project = str(row.get('Project/Group', '')).strip()

            # Normalize invalid or missing values
            if person.lower() in ['nan', 'none', '']:
                person = ''
            if project.lower() in ['nan', 'none', '']:
                project = ''

            # Assign a unique color to a person if not already assigned
            if person:
                if person not in assigned_person_colors:
                    assigned_person_colors[person] = get_next_color()
                return PatternFill(start_color=assigned_person_colors[person], end_color=assigned_person_colors[person], fill_type="solid")
            # Otherwise assign a unique color to a project if not already assigned
            elif project:
                if project not in assigned_project_colors:
                    assigned_project_colors[project] = get_next_color()
                return PatternFill(start_color=assigned_project_colors[project], end_color=assigned_project_colors[project], fill_type="solid")
            # Fill in gray as defualt if no person or project is defined
            else:
                return PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Iterate over each row in the DataFrame to update grid positions
        for _, row in self.rows_df.iterrows():
            try:
                # Read shelf and rack numbers from the DataFrame
                shelf_num = row['Shelf Number']
                rack_num = row['Rack Number']
                # Skip rows with invalid shelf or rack numbers
                if not (1 <= shelf_num <= 5 and 1 <= rack_num <= 4):
                    continue

                # Get box position (e.g., "A1") and box label
                position = str(row['Box Position']).strip().upper()
                label = str(row['Box Name']).strip()
                if len(position) < 2:
                    continue

                # Split position into row letter and column number
                row_letter = position[0]
                col_number = position[1:]

                # Validate row letter (A–D)
                if row_letter not in "ABCD":
                    continue
                
                # Validate column number (1–5)
                if not col_number.isdigit():
                    continue
                col_number = int(col_number)
                if not (1 <= col_number <= 5):
                    continue

                # Calculate row and column offsets
                row_offset = ord(row_letter) - ord('A')   # A=0, B=1, C=2, D=3
                col_offset = int(col_number) - 1          # 1=0, …, 5=4
                shelf_offset = (shelf_num - 1) * 8        # Shelf offset: Shelf 1 starts at row 3, Shelf 2 at 11, Shelf 3 at 19, etc.

                # Compute final Excel row index. Rack base row: Row A = Excel row 3 (inside shelf)
                row_index = 3 + row_offset + shelf_offset

                # Compute final Excel column index. Rack base column: rack 1 starts at C (3), rack 2 at J (10), rack 3 at Q (17), rack 4 at X (24)
                rack_start_col = 3 + (rack_num - 1) * 7
                col_index = rack_start_col + col_offset

                # Get the target cell
                cell = sheet.cell(row=row_index, column=col_index)
                cell.value = label
                # Align text in center, wrap long text
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                # Assign fill color using get_fill_color()
                cell.fill = get_fill_color(row)

            except Exception as e:
                # Print error but continue processing other rows
                print(f"Error processing row {row}: {e}")

        # Clear legend area (columns 31–32) for 50 rows to prepare for fresh legend
        for row in range(1, 51):
            for col in range(31, 33):
                cell = sheet.cell(row=row, column=col)
                cell.value = None
                cell.fill = PatternFill()

        # Write legend title in column 31
        legend_col = 31
        legend_row = 1
        sheet.cell(row=legend_row, column=legend_col, value="Color Legend").font = Font(bold=True)
        legend_row += 2  # Start legend entries two rows below title

        # Add person legend entries, sorted alphabetically
        for person, color in sorted(assigned_person_colors.items()):
            cell = sheet.cell(row=legend_row, column=legend_col)
            cell.value = f"Person: {person}"
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            legend_row += 1

        # Add project legend entries, sorted alphabetically
        for project, color in sorted(assigned_project_colors.items()):
            cell = sheet.cell(row=legend_row, column=legend_col)
            cell.value = f"Project: {project}"
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            legend_row += 1

        # Save workbook after updates
        wb.save(self.grid_path)
        print(f"✅ -80 Freezer inventory updated and saved to: {self.grid_path}")

class Freezer20Manager(InventoryManagerBase):
    """
    Manages inventory records for the -20 Freezer system.

    This subclass of InventoryManagerBase provides configuration details
    specific to -20 Freezer storage, including file paths, required fields,
    sorting rules, and user interface labels. It defines how rows should
    be validated, displayed, and managed for entries in the -20 Freezer
    inventory, and handles updating the Excel-based grid layout with
    colored indicators for person/project assignments.

    Parameters
    ----------
    root (tk.Widget):
        The Tkinter root window or parent widget.

    Attributes
    ----------
    row_id_key (str):
        Key used to look up the file ID for row-based inventory in the ID manager.
    grid_id_key (str or None):
        Key used to look up the file ID for grid-based inventory in the ID manager.
    row_ID (str):
        File ID for the row-based inventory Excel file.
    grid_ID (str or None):
        File ID for grid-based inventory Excel file.

    Methods
    -------
    get_window_title() -> str:
        Return the title of the main inventory window.
    get_row_sheet_name() -> str:
        Return the name of the Excel sheet used for row-based inventory.
    get_sort_columns() -> list[str]:
        Return the columns used to sort the -20 Freezer inventory.
    get_for_filled() -> list[str]:
        Return the columns used to determine if an inventory position is filled.
    get_location_column() -> str:
        Return the column representing the physical location of an inventory unit.
    get_int_fields() -> list[str]:
        Return the fields that should be strictly treated as integers.
    get_remove_fields() -> list[str]:
        Return the fields required to identify and remove an inventory entry.
    get_label_column() -> str:
        Return the column used as a label for an inventory entry.
    get_letter_nums() -> list[str]:
        Return the fields that contain a letter followed by a number (e.g., A4).
    get_grid_coords() -> tuple[int, int]:
        Return the dimensions (rows, columns) for the picker pop-up grid.
    get_required_add_fields() -> list[str]:
        Return the fields required when adding a new entry.
    get_add_top_label() -> str:
        Return the label shown at the top of the 'Add Entry' window.
    get_add_window_name() -> str:
        Return the name of the 'Add Entry' window.
    get_required_remove_fields() -> list[str]:
        Return the fields required when removing an entry.
    get_remove_top_label() -> str:
        Return the label shown at the top of the 'Remove Entry' window.
    get_remove_window_name() -> str:
        Return the name of the 'Remove Entry' window.
    get_unused_columns() -> list[str]:
        Return column names that should be ignored in the GUI.
    update_grid_inventory() -> None:
        Update the Excel grid layout for the -20 Freezer, clearing existing values,
        placing boxes in the correct shelf/rack/position, assigning colors to persons/projects,
        creating a legend, saving the workbook, and uploading it to Google Drive.
    """
            
    def __init__(self, root):
        """Initializes the Freezer20Manager. See class docstring for parameter/attribute details."""
        
        # Initialize logic from the InventoryManagerBase using the given row_path and grid_path
        super().__init__(
            root,
            row_path="20_Inventory.xlsx",
            grid_path="20_Inventory.xlsx"
        )
        # Key for row inventory lookup in the ID manager
        self.row_id_key = "20_Inventory"
        # Key for grid inventory lookup in the ID manager
        self.grid_id_key = "20_Inventory"
        # Get the unique ID for the row file from the ID manager
        self.row_ID = app_context.id_manager.get_id(self.row_id_key)
        # Get the unique ID for the grid file from the ID manager
        self.grid_ID = app_context.id_manager.get_id(self.grid_id_key)

    def get_window_title(self):
        """Return the title for the main window."""

        return "-20 Freezer Dewar"

    def get_row_sheet_name(self):
        """Return the name of the Excel sheet used for row-based inventory."""

        return "Details"

    def get_sort_columns(self):
        """Return the list of columns used for sorting the grid dewar inventory in sorting order."""

        return ["Shelf Number", "Rack Number", "Box Position"]

    def get_for_filled(self):
        """
        Return the columns used to determine if an inventory position is filled.

        This excludes the column containing the location of an individual unit of inventory.
        """
                
        return ["Shelf Number", "Rack Number"]

    def get_location_column(self):
        """Return the column representing the physical location of an individual unit of inventory."""

        return "Box Position"

    def get_int_fields(self):
        """Return the fields that should be strictly treated as integers."""

        return ["Shelf Number", "Rack Number"]

    def get_remove_fields(self):
        """Return the fields required to identify and remove an entry."""

        return ["Shelf Number", "Rack Number", "Box Position"]

    def get_label_column(self):
        """Return the column name used as a label for an individual inventory entry."""

        return "Box Name"

    def get_letter_nums(self):
        """Return the fields that conatin a letter followed by a number (Ex: A4)"""

        return ["Box Position"]
    
    def get_grid_coords(self):
        """Return the dimensions for the picker pop-up (rows, columns)."""

        return (5, 3)

    def get_required_add_fields(self):
        """Return the fields required to be enetered by the user when adding a new entry."""

        return ["Shelf Number", "Rack Number", "Box Position", "Box Name", "Person/Initials", "Project/Group"]

    def get_add_top_label(self):
        """Return the label shown at the top of the 'Add Entry' window."""

        return "Enter Box Information Below"

    def get_add_window_name(self):
        """Return the name of the 'Add Entry' window."""

        return "-20 Freezer Box Entry"

    def get_required_remove_fields(self):
        """Return the fields to be enterd by the user required when removing an entry."""

        return ["Shelf Number", "Rack Number", "Box Position"]

    def get_remove_top_label(self):
        """Return the label shown at the top of the 'Remove Entry' window."""

        return "Enter Box Information Below"

    def get_remove_window_name(self):
        """Return the name of the 'Remove Entry' window."""

        return "-20 Freezer Box Removal"

    def get_unused_columns(self):
        """Return column names that are ignored in the GUI."""

        return []

    def update_grid_inventory(self):
        """
        Update the Excel grid inventory layout for the -20°C freezer.

        This method clears existing entries in the predefined grid layout
        and fills in the updated box names from `self.rows_df`. Each box
        is positioned based on its shelf, rack, and position information. Cells
        are automatically colored based on the person or project responsible,
        with a color legend generated on the side of the sheet.

        Steps performed:
        1. Clear all values, formatting, and alignment in the defined row/column blocks.
        2. Define a pool of colors for assigning to persons or projects.
        3. Assign colors dynamically to unique persons/projects as boxes are added.
        4. Validate and place each box in the correct shelf/rack/position.
        5. Create a legend on the side showing assigned colors for persons/projects.
        6. Save the Excel workbook.

        Notes
        -----
        - Expects the Excel workbook to have a "Racks" sheet.
        - Shelf numbers should be 1–5, racks 1–4, row letters A–E, columns 1–3.
        - Any invalid or missing positions are skipped.
        - Fallback color "D3D3D3" is used if color pool is exhausted.

        Methods
        -------
        def get_next_color():
            Return the next unused color from the color pool.
        get_fill_color(row):
            Determine cell fill color based on 'Person/Initials' or 'Project/Group'.
        """

        # Load the workbook and the "Racks" sheet
        wb = load_workbook(self.grid_path)
        sheet = wb["Racks"]

        # Define the row blocks to clear before updating
        row_blocks = [
            range(3, 8),    # Rows 3–7
            range(12, 17),  # Rows 12-16
            range(21, 26),  # Rows 21-25
            range(30, 35),  # Rows 30-34
            range(39, 44),  # Rows 39–43
        ]

        # Define the column groups to clear before updating
        col_groups = [
            ("C", "E"),   # Columns C–E
            ("H", "J"),   # Columns H–J
            ("M", "O"),   # Columns M–O
            ("R", "T"),  # Columns R–T
        ]

        # Clear existing values, fills, and alignments
        for row_block in row_blocks:
            for start_col, end_col in col_groups:
                start_idx = column_index_from_string(start_col)
                end_idx = column_index_from_string(end_col)
                for r in row_block:
                    for c in range(start_idx, end_idx + 1):
                        cell = sheet.cell(row=r, column=c)
                        cell.value = ""  # clear value
                        cell.fill = PatternFill()  # clear fill
                        cell.alignment = Alignment(
                            wrap_text=True, horizontal="center", vertical="center"
                        )

        # Define a pool of colors for person/project assignments
        color_pool = [
            "FFB3BA", "FFB74D", "FFFFBA", "A5D6A7", "81D4FA", "B39DDB",
            "E53935", "FB8C00", "FDD835", "43A047", "039BE5", "00ACC1", "8E24AA",
            "D7CCC8", "FFE0B2", "8E735B", "6D4C41", "8D6E63",
            "BDBDBD", "757575", "90A4AE", "A5A58D", "546E7A",
            "00897B", "7CB342", "CE93D8", "AED581", "FFD54F", "FF8A65", "C0CA33",
            "FF1744", "D500F9", "00E5FF", "1DE9B6", "FFD600", "FF6D00", "DD2C00",
            "304FFE", "00BFA5", "76FF03", "C51162", "6200EA", "2962FF", "00B0FF",
            "00C853", "FFD600", "FFAB00", "FF3D00", "B71C1C", "1B5E20", "0D47A1",
            "311B92", "F57F17", "FF1744", "D50000", "C2185B", "4A148C", "0091EA",
            "00BFA5", "64DD17", "FFD600"
        ]

        # Track which colors are assigned to which persons/projects
        assigned_person_colors = {}
        assigned_project_colors = {}
        used_colors = set()

        def get_next_color():
            """Return the next unused color from the color pool."""

            for color in color_pool:
                if color not in used_colors:
                    used_colors.add(color)
                    return color
            return "D3D3D3"  # fallback gray

        def get_fill_color(row):
            """
            Determine cell fill color based on 'Person/Initials' or 'Project/Group'.
            
            Priority is given to person. If neither is defined, a default gray is used.
            """
            # Get the 'Person/Initials" and "Project/Group" of the row
            person = str(row.get('Person/Initials', '')).strip()
            project = str(row.get('Project/Group', '')).strip()

            # Normalize invalid or missing values
            if person.lower() in ['nan', 'none', '']:
                person = ''
            if project.lower() in ['nan', 'none', '']:
                project = ''

            # Assign a unique color to a person if not already assigned
            if person:
                if person not in assigned_person_colors:
                    assigned_person_colors[person] = get_next_color()
                return PatternFill(start_color=assigned_person_colors[person], end_color=assigned_person_colors[person], fill_type="solid")
            # Otherwise assign a unique color to a project if not already assigned
            elif project:
                if project not in assigned_project_colors:
                    assigned_project_colors[project] = get_next_color()
                return PatternFill(start_color=assigned_project_colors[project], end_color=assigned_project_colors[project], fill_type="solid")
            # Fill in gray as defualt if no person or project is defined
            else:
                return PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Iterate over each row in the DataFrame to update grid positions
        for _, row in self.rows_df.iterrows():
            try:
                # Read shelf and rack numbers from the DataFrame
                shelf_num = row['Shelf Number']
                rack_num = row['Rack Number']
                # Skip rows with invalid shelf or rack numbers
                if not (1 <= shelf_num <= 5 and 1 <= rack_num <= 4):
                    continue

                # Get box position (e.g., "A1") and box label
                position = str(row['Box Position']).strip().upper()
                label = str(row['Box Name']).strip()
                if len(position) < 2:
                    continue

                # Split position into row letter and column number
                row_letter = position[0]
                col_number = position[1:]

                # Validate row letter (A–E)
                if row_letter not in "ABCDE":
                    continue
                # Validate column number (1–3)
                if not col_number.isdigit():
                    continue
                col_number = int(col_number)
                if not (1 <= col_number <= 3):
                    continue

                # Calculate row and column offsets
                row_offset = ord(row_letter) - ord('A')   # A=0, B=1, C=2, D=3, E=4
                col_offset = int(col_number) - 1          # 1=0, 2=1, 3=2
                shelf_offset = (shelf_num - 1) * 9        # Shelf offset: Shelf 1 starts at row 3, Shelf 2 at 11, Shelf 3 at 19, etc.

                # Compute final Excel row index. Rack base row: Row A = Excel row 3 (inside shelf)
                row_index = 3 + row_offset + shelf_offset

                # Compute final Excel column index. Rack base column: rack 1 starts at C (3), rack 2 at J (10), rack 3 at Q (17), rack 4 at X (24)
                rack_start_col = 3 + (rack_num - 1) * 5
                col_index = rack_start_col + col_offset

                # Get the target cell
                cell = sheet.cell(row=row_index, column=col_index)
                cell.value = label
                # Align text in center, wrap long text
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                # Assign fill color using get_fill_color()
                cell.fill = get_fill_color(row)
            
            except Exception as e:
                # Print error but continue processing other rows
                print(f"Error processing row {row}: {e}")

        # Clear legend area (columns 23-24) for 50 rows to prepare for fresh legend
        for row in range(1, 51):
            for col in range(23, 25):
                cell = sheet.cell(row=row, column=col)
                cell.value = None
                cell.fill = PatternFill()

        # Write legend title in column 23
        legend_col = 23
        legend_row = 1
        sheet.cell(row=legend_row, column=legend_col, value="Color Legend").font = Font(bold=True)
        legend_row += 2  # Start legend entries two rows below title

        # Add person legend entries, sorted alphabetically
        for person, color in sorted(assigned_person_colors.items()):
            cell = sheet.cell(row=legend_row, column=legend_col)
            cell.value = f"Person: {person}"
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            legend_row += 1

        # Add project legend entries, sorted alphabetically
        for project, color in sorted(assigned_project_colors.items()):
            cell = sheet.cell(row=legend_row, column=legend_col)
            cell.value = f"Project: {project}"
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
            legend_row += 1

        # Save workbook after updates
        wb.save(self.grid_path)
        print(f"✅ -20 Freezer inventory updated and saved to: {self.grid_path}")

class CellDewarManager(InventoryManagerBase):
    """
    Manages inventory records for the Cell Dewar system.

    This subclass of InventoryManagerBase provides configuration details
    specific to cell dewar storage, including file paths, required fields,
    sorting rules, and user interface labels. It defines how rows should
    be validated, displayed, and managed for entries in the cell dewar
    inventory, and handles updating the Excel-based grid layout with vial
    labels and standardized formatting.

    Parameters
    ----------
    root (tk.Widget): 
        The Tkinter root window or parent widget.

    Attributes
    ----------
    row_id_key (str):
        Key used to look up the file ID for row-based inventory in the ID manager.
    grid_id_key (str or None):
        Key used to look up the file ID for grid-based inventory in the ID manager.
    row_ID (str):
        File ID for the row-based inventory Excel file.
    grid_ID (str or None):
        File ID for grid-based inventory Excel file.

    Methods
    -------
    get_window_title() -> str:
        Return the title of the main inventory window.
    get_row_sheet_name() -> str:
        Return the name of the Excel sheet used for row-based inventory.
    get_sort_columns() -> list[str]:
        Return the columns used to sort the cell dewar inventory.
    get_for_filled() -> list[str]:
        Return the columns used to determine if an inventory position is filled.
    get_location_column() -> str:
        Return the column representing the physical location of an inventory unit.
    get_int_fields() -> list[str]:
        Return the fields that should be strictly treated as integers.
    get_remove_fields() -> list[str]:
        Return the fields required to identify and remove an inventory entry.
    get_label_column() -> str:
        Return the column used as a label for an inventory entry.
    get_letter_nums() -> list[str]:
        Return the fields that contain a letter followed by a number (e.g., A4).
    get_date_column() -> str:
        Return the column storing date information.
    get_grid_coords() -> tuple[int, int]:
        Return the dimensions (rows, columns) for the picker pop-up grid.
    get_required_add_fields() -> list[str]:
        Return the fields required when adding a new entry.
    get_add_top_label() -> str:
        Return the label shown at the top of the 'Add Entry' window.
    get_add_window_name() -> str:
        Return the name of the 'Add Entry' window.
    get_required_remove_fields() -> list[str]:
        Return the fields required when removing an entry.
    get_remove_top_label() -> str:
        Return the label shown at the top of the 'Remove Entry' window.
    get_remove_window_name() -> str:
        Return the name of the 'Remove Entry' window.
    get_unused_columns() -> list[str]:
        Return column names that should be ignored in the GUI.
    update_grid_inventory() -> None:
        Update the Excel grid layout for cell dewar inventory.
    """

    def __init__(self, root):
        """Initializes the Freezer80Manager. See class docstring for parameter/attribute details."""

        # Initialize logic from the InventoryManagerBase using the given row_path and grid_path
        super().__init__(
            root,
            row_path="Cell_Culture_Inventory_Rows.xlsx",
            grid_path="Cell_Culture_Inventory_Grid.xlsx"
        )
        # Key for row inventory lookup in the ID manager
        self.row_id_key = "Cell_Culture_Inventory_Rows"
        # Key for grid inventory lookup in the ID manager
        self.grid_id_key = "Cell_Culture_Inventory_Grid"
        # Get the unique ID for the row file from the ID manager
        self.row_ID = app_context.id_manager.get_id(self.row_id_key)
        # Get the unique ID for the grid file from the ID manager
        self.grid_ID = app_context.id_manager.get_id(self.grid_id_key)

    def get_window_title(self):
        """Return the title for the main window."""

        return "Cell Dewar"

    def get_row_sheet_name(self):
        """Return the name of the Excel sheet used for row-based inventory."""

        return "Details"

    def get_sort_columns(self):
        """Return the list of columns used for sorting the grid dewar inventory in sorting order."""

        return ["Rack Number", "Box Number", "Vial Position"]

    def get_for_filled(self):
        """
        Return the columns used to determine if an inventory position is filled.

        This excludes the column containing the location of an individual unit of inventory.
        """

        return ["Rack Number", "Box Number"]

    def get_location_column(self):
        """Return the column representing the physical location of an individual unit of inventory."""

        return "Vial Position"

    def get_int_fields(self):
        """Return the fields that should be strictly treated as integers."""

        return ["Rack Number", "Box Number"]

    def get_remove_fields(self):
        """Return the fields required to identify and remove an entry."""

        return ["Rack Number", "Box Number", "Vial Position"]

    def get_label_column(self):
        """Return the column name used as a label for an individual inventory entry."""

        return "Vial Label"

    def get_letter_nums(self):
        """Return the fields that conatin a letter followed by a number (Ex: A4)"""

        return ["Vial Position", "Passage Number"]

    def get_date_column(self):
        """Return the column that stores date information."""

        return "Date Frozen"

    def get_grid_coords(self):
        """Return the dimensions for the picker pop-up (rows, columns)."""

        return (9, 9)

    def get_required_add_fields(self):
        """Return the fields required to be enetered by the user when adding a new entry."""

        return ["Rack Number", "Box Number", "Vial Position", "Vial Label", "Cell Type", "Passage Number", "Date Frozen", "Person/Initials", "Project"]

    def get_add_top_label(self):
        """Return the label shown at the top of the 'Add Entry' window."""

        return "Enter Vial Information Below"

    def get_add_window_name(self):
        """Return the name of the 'Add Entry' window."""

        return "Cell Culture Dewar Vial Entry"

    def get_required_remove_fields(self):
        """Return the fields to be enterd by the user required when removing an entry."""

        return ["Rack Number", "Box Number", "Vial Position"]

    def get_remove_top_label(self):
        """Return the label shown at the top of the 'Remove Entry' window."""

        return "Enter Vial Information Below"

    def get_remove_window_name(self):
        """Return the name of the 'Remove Entry' window."""

        return "Cell Culture Dewar Vial Removal"

    def get_unused_columns(self):
        """Return column names that are ignored in the GUI."""

        return ['Unnamed: 4', 'Original Box']

    def update_grid_inventory(self):
        """
        Update the Excel grid inventory with vial labels and standardized formatting.

        This function reads inventory information from `self.rows_df` and writes it
        to an Excel workbook located at `self.grid_path`. Each sheet in the workbook
        corresponds to a box and is labeled the box number, and vial positions are
        mapped to specific cells based on row letters (A-I) and column numbers (1-9).

        Notes:
        - Only valid row letters (A-I) and column numbers (1-9) are processed.
        - Empty or invalid positions are skipped with a warning printed to the console.
        - Each vial label is centered and wrapped in its cell for readability.
        - All sheets are cleared in the target range before writing new values to prevent stale data.
        """
        
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Load the Excel workbook containing the grid box inventory
        wb = load_workbook(self.grid_path)

        # Create a copy of the DataFrame to prevent changes
        source_df = self.rows_df.copy()
        source_df.columns = [col.strip() for col in source_df.columns]

        # Map rows A-I to excel rows 2-10
        row_map = {chr(i): i - 63 for i in range(65, 74)}  # A=2,... I=10
        # Map cols 1-9 to excel columns B-J
        col_map = {str(i): get_column_letter(i + 1) for i in range(1, 10)}

        # Clear all sheets in the workbook
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Clears only in rows 2-10 and columns 2-10 where vials will be placed
            for row in range(2, 11):
                for col in range(2, 11):
                    cell = sheet.cell(row=row, column=col)
                    # Remove text
                    cell.value = ""
                    # Remove color fill
                    cell.fill = PatternFill()
                    # Reset alignment: centered and wrap text
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Iterate through each row in the DataFrame to update cells in the grid inventory sheet
        for _, row in source_df.iterrows():
            # Skip rows without valid box numbers
            box_val = row['Box Number']
            if box_val is None or (isinstance(box_val, float) and math.isnan(box_val)):
                print(f"⚠️ Skipping row with invalid Box: {box_val}")
                continue

            # Convert float box numbers to integers, otherwise just strip string
            box = str(int(row['Box Number'])) if isinstance(row['Box Number'], float) else str(row['Box Number']).strip()
            position = str(row['Vial Position']).strip().upper()
            label = str(row['Vial Label']).strip()

            # Force datetime into string if needed
            if isinstance(row['Vial Label'], pd.Timestamp):
                label = row['Vial Label'].strftime("%m/%d/%Y")

            # Skip invalid or too short positions
            if len(position) < 2:
                continue

            # Split the position to get the row (letter) and column (number)
            row_letter = position[0]
            col_number = position[1:]

            # Skip positions outside of mapped ranges
            if row_letter not in row_map or col_number not in col_map:
                print(f"⚠️ Skipping invalid position: {position}")
                continue

            # Convert the given position to the position of its label in excel
            row_index = row_map[row_letter]
            col_letter = col_map[col_number]
            cell_ref = f"{col_letter}{row_index}"

            # Only update the sheet if it exists for this box number
            if box in wb.sheetnames:
                sheet = wb[box]
                cell = sheet[cell_ref]
                cell.value = label  # Write the vial label
                # Wrap text and center for readability
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.fill = gray_fill  # Fill occupied cell with gray
            else:
                print(f"❌ Sheet '{box}' not found in workbook")

        # Save the workbook after all updates
        wb.save(self.grid_path)
        print(f"✅ Done! Saved to {self.grid_path}")